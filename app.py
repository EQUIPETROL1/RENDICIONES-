import os, json, base64, re, uuid
from datetime import datetime
from pathlib import Path
from flask import Flask, request, jsonify, render_template, send_file, session, redirect, url_for
import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "rendiciones-secret-2024")

BASE_DIR = Path(__file__).parent
FACTURAS_DIR = BASE_DIR / "facturas"
RENDICIONES_DIR = BASE_DIR / "rendiciones_excel"
WORKERS_FILE = BASE_DIR / "workers.json"
FACTURAS_DIR.mkdir(exist_ok=True)
RENDICIONES_DIR.mkdir(exist_ok=True)

ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")

def load_workers():
    if WORKERS_FILE.exists():
        return json.loads(WORKERS_FILE.read_text())
    return {}

def save_workers(data):
    WORKERS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2))

def load_worker_invoices(worker_id):
    path = FACTURAS_DIR / worker_id / "invoices.json"
    if path.exists():
        return json.loads(path.read_text())
    return []

def save_worker_invoices(worker_id, invoices):
    d = FACTURAS_DIR / worker_id
    d.mkdir(exist_ok=True)
    (d / "invoices.json").write_text(json.dumps(invoices, ensure_ascii=False, indent=2))

def extract_invoice_data(file_bytes, mime_type, filename):
    genai.configure(api_key=os.environ.get("GEMINI_API_KEY"))
    model = genai.GenerativeModel("gemini-1.5-flash")
    prompt = (
        "Extrae los datos de esta factura. Responde SOLO con JSON válido, sin markdown ni texto extra. "
        "Usa exactamente estos campos: "
        '{"numero_factura":"","fecha_emision":"","fecha_vencimiento":"","proveedor":"","ruc_proveedor":"",'
        '"cliente":"","ruc_cliente":"","descripcion":"","base_gravada":"","igv":"","total":"","moneda":"","forma_pago":""}. '
        "Para montos usa solo el número (ej: 45.25). Si un campo no existe déjalo vacío."
    )
    part = {"mime_type": mime_type, "data": file_bytes}
    response = model.generate_content([prompt, part])
    text = response.text
    text = re.sub(r"```json|```", "", text).strip()
    data = json.loads(text)
    data["archivo"] = filename
    data["fecha_carga"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    return data

def generate_excel(worker_id, worker_name, invoices):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rendición"
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.merge_cells("A1:M1")
    t = ws["A1"]
    t.value = f"RENDICIÓN DE GASTOS — {worker_name.upper()}"
    t.font = Font(name="Arial", bold=True, size=13, color="1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:M2")
    s = ws["A2"]
    s.value = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    s.font = Font(name="Arial", size=9, color="888888")
    s.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16
    headers = ["N° Factura","Fecha Emisión","Fecha Venc.","Proveedor","RUC Prov.",
               "Cliente","RUC Cliente","Descripción","Base Gravada","IGV","Total","Moneda","Forma Pago"]
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = border
    ws.row_dimensions[4].height = 30
    fills = [PatternFill("solid", fgColor="FFFFFF"), PatternFill("solid", fgColor="EBF5FB")]
    dfont = Font(name="Arial", size=10)
    for i, inv in enumerate(invoices):
        r = 5 + i
        row_data = [
            inv.get("numero_factura",""), inv.get("fecha_emision",""), inv.get("fecha_vencimiento",""),
            inv.get("proveedor",""), inv.get("ruc_proveedor",""), inv.get("cliente",""),
            inv.get("ruc_cliente",""), inv.get("descripcion",""),
            _num(inv.get("base_gravada","")), _num(inv.get("igv","")), _num(inv.get("total","")),
            inv.get("moneda",""), inv.get("forma_pago","")
        ]
        fill = fills[i % 2]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = dfont; c.fill = fill; c.border = border
            if ci in (9, 10, 11):
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif ci in (1,2,3,5,7,12,13):
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 20
    tr = 5 + len(invoices)
    tfill = PatternFill("solid", fgColor="1F4E79")
    tfont = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    for ci in range(1, 14):
        c = ws.cell(row=tr, column=ci)
        c.fill = tfill; c.font = tfont; c.border = border
    ws.cell(row=tr, column=8, value="TOTALES").alignment = Alignment(horizontal="right", vertical="center")
    for ci, col in [(9,"I"),(10,"J"),(11,"K")]:
        c = ws.cell(row=tr, column=ci, value=f"=SUM({col}5:{col}{tr-1})")
        c.number_format = "#,##0.00"
        c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[tr].height = 24
    col_widths = [16,13,13,36,13,28,13,32,14,11,11,10,12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    out = RENDICIONES_DIR / f"Rendicion_{worker_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out)
    return out

def _num(val):
    try:
        return float(str(val).replace(",","").strip())
    except:
        return val

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/register", methods=["POST"])
def register():
    data = request.json
    name = data.get("name","").strip()
    if not name:
        return jsonify({"error": "Nombre requerido"}), 400
    workers = load_workers()
    for wid, info in workers.items():
        if info["name"].lower() == name.lower():
            session["worker_id"] = wid
            session["worker_name"] = info["name"]
            return jsonify({"worker_id": wid, "name": info["name"], "existing": True})
    wid = str(uuid.uuid4())[:8]
    workers[wid] = {"name": name, "created": datetime.now().isoformat()}
    save_workers(workers)
    session["worker_id"] = wid
    session["worker_name"] = name
    return jsonify({"worker_id": wid, "name": name, "existing": False})

@app.route("/api/upload", methods=["POST"])
def upload():
    worker_id = request.form.get("worker_id") or session.get("worker_id")
    if not worker_id:
        return jsonify({"error": "Sin sesión"}), 401
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No se recibió archivo"}), 400
    allowed = {"application/pdf","image/jpeg","image/png","image/jpg"}
    if file.mimetype not in allowed:
        return jsonify({"error": f"Formato no soportado: {file.mimetype}"}), 400
    file_bytes = file.read()
    try:
        inv_data = extract_invoice_data(file_bytes, file.mimetype, file.filename)
    except Exception as e:
        return jsonify({"error": f"Error al extraer: {str(e)}"}), 500
    worker_dir = FACTURAS_DIR / worker_id
    worker_dir.mkdir(exist_ok=True)
    safe_name = f"{uuid.uuid4().hex[:8]}_{file.filename}"
    (worker_dir / safe_name).write_bytes(file_bytes)
    inv_data["_file_saved"] = safe_name
    invoices = load_worker_invoices(worker_id)
    inv_data["_id"] = str(uuid.uuid4())[:8]
    invoices.append(inv_data)
    save_worker_invoices(worker_id, invoices)
    return jsonify({"ok": True, "invoice": inv_data})

@app.route("/api/invoices/<worker_id>")
def get_invoices(worker_id):
    return jsonify(load_worker_invoices(worker_id))

@app.route("/api/update_invoice", methods=["POST"])
def update_invoice():
    data = request.json
    worker_id = data.get("worker_id") or session.get("worker_id")
    inv_id = data.get("inv_id")
    field = data.get("field")
    value = data.get("value")
    invoices = load_worker_invoices(worker_id)
    for inv in invoices:
        if inv.get("_id") == inv_id:
            inv[field] = value
            break
    save_worker_invoices(worker_id, invoices)
    return jsonify({"ok": True})

@app.route("/api/delete_invoice", methods=["POST"])
def delete_invoice():
    data = request.json
    worker_id = data.get("worker_id") or session.get("worker_id")
    inv_id = data.get("inv_id")
    invoices = load_worker_invoices(worker_id)
    invoices = [i for i in invoices if i.get("_id") != inv_id]
    save_worker_invoices(worker_id, invoices)
    return jsonify({"ok": True})

@app.route("/api/export/<worker_id>")
def export(worker_id):
    workers = load_workers()
    worker_name = workers.get(worker_id, {}).get("name", worker_id)
    invoices = load_worker_invoices(worker_id)
    if not invoices:
        return jsonify({"error": "Sin facturas"}), 400
    path = generate_excel(worker_id, worker_name, invoices)
    return send_file(path, as_attachment=True, download_name=path.name)

@app.route("/admin", methods=["GET","POST"])
def admin():
    if request.method == "POST":
        pw = request.form.get("password","")
        if pw == ADMIN_PASSWORD:
            session["is_admin"] = True
            return redirect(url_for("admin"))
        return render_template("admin_login.html", error="Contraseña incorrecta")
    if not session.get("is_admin"):
        return render_template("admin_login.html", error=None)
    workers = load_workers()
    summary = []
    for wid, info in workers.items():
        invs = load_worker_invoices(wid)
        total = sum(_num(i.get("total","0")) for i in invs if i.get("total"))
        try: total = float(total)
        except: total = 0
        summary.append({"id": wid, "name": info["name"], "count": len(invs), "total": total})
    return render_template("admin.html", workers=summary)

@app.route("/api/admin/export_all")
def export_all():
    if not session.get("is_admin"):
        return jsonify({"error": "No autorizado"}), 403
    workers = load_workers()
    wb = Workbook()
    first = True
    for wid, info in workers.items():
        invs = load_worker_invoices(wid)
        if not invs:
            continue
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = info["name"][:31]
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.merge_cells("A1:M1")
        t = ws["A1"]
        t.value = f"RENDICIÓN — {info['name'].upper()}"
        t.font = Font(name="Arial", bold=True, size=12, color="1F4E79")
        t.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 26
        headers = ["N° Factura","Fecha","Proveedor","RUC Prov.","Cliente","RUC Cliente",
                   "Descripción","Base Gravada","IGV","Total","Moneda","Forma Pago","Archivo"]
        hfill = PatternFill("solid", fgColor="1F4E79")
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=ci, value=h)
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            c.fill = hfill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        ws.row_dimensions[3].height = 28
        fills = [PatternFill("solid", fgColor="FFFFFF"), PatternFill("solid", fgColor="EBF5FB")]
        for i, inv in enumerate(invs):
            r = 4 + i
            row_data = [inv.get("numero_factura",""), inv.get("fecha_emision",""),
                        inv.get("proveedor",""), inv.get("ruc_proveedor",""),
                        inv.get("cliente",""), inv.get("ruc_cliente",""),
                        inv.get("descripcion",""), _num(inv.get("base_gravada","")),
                        _num(inv.get("igv","")), _num(inv.get("total","")),
                        inv.get("moneda",""), inv.get("forma_pago",""), inv.get("archivo","")]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=r, column=ci, value=val)
                c.font = Font(name="Arial", size=10)
                c.fill = fills[i%2]; c.border = border
                if ci in (8,9,10):
                    c.number_format = "#,##0.00"
                    c.alignment = Alignment(horizontal="right")
                else:
                    c.alignment = Alignment(horizontal="left", wrap_text=True)
            ws.row_dimensions[r].height = 18
        tr = 4 + len(invs)
        for ci in range(1, 14):
            c = ws.cell(row=tr, column=ci)
            c.fill = PatternFill("solid", fgColor="1F4E79")
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            c.border = border
        ws.cell(row=tr, column=7, value="TOTALES").alignment = Alignment(horizontal="right")
        for ci, col in [(8,"H"),(9,"I"),(10,"J")]:
            c = ws.cell(row=tr, column=ci, value=f"=SUM({col}4:{col}{tr-1})")
            c.number_format = "#,##0.00"
            c.alignment = Alignment(horizontal="right")
        ws.row_dimensions[tr].height = 22
        col_widths = [16,13,30,13,26,13,30,13,11,11,10,12,20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A4"
    if first:
        return jsonify({"error": "Sin datos"}), 400
    fname = RENDICIONES_DIR / f"Rendiciones_Todas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(fname)
    return send_file(fname, as_attachment=True, download_name=fname.name)

@app.route("/api/admin/workers")
def admin_workers():
    if not session.get("is_admin"):
        return jsonify({"error": "No autorizado"}), 403
    workers = load_workers()
    result = []
    for wid, info in workers.items():
        invs = load_worker_invoices(wid)
        total = sum((_num(i.get("total","0")) or 0) for i in invs)
        result.append({"id": wid, "name": info["name"], "count": len(invs), "total": round(float(total),2)})
    return jsonify(result)

if __name__ == "__main__":
    app.run(debug=True, port=5000)
